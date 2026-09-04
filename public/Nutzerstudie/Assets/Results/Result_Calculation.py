#!/usr/bin/env python3
"""
Auswertung der Nutzerstudie "Visualisierung von Unsicherheiten in 2D-Daten"

Nutzung:
- Diese .py-Datei in denselben Ordner wie die Ergebnis-CSV legen.
- In Visual Studio Code öffnen und ausführen.
- Standardmäßig wird automatisch die erste passende CSV im selben Ordner gesucht.
- Die Ergebnisse werden als CSV und zusätzlich gesammelt als XLSX im Unterordner "analysis_output" gespeichert.

WICHTIG:
- Antwort -999 = nicht beantwortet.
- -999 wird als fehlende Antwort behandelt und NICHT als Ausreißer oder Fehlerwert.
- Es findet keine automatische Ausreißerentfernung statt.
"""

from __future__ import annotations

import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats
from statsmodels.stats.anova import AnovaRM


# ---------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------

METHOD_ORDER = ["VSUP", "ScaledGlyph", "IsoGlyph"]

# Feste Darstellungsreihenfolge in der Excel-Auswertung.
# Die tatsächliche Bearbeitungsreihenfolge bleibt zusätzlich in trialOrder erhalten.
H1_TASK_ORDER = [
    "Niederschlag_Uncertainty_Maxima",
    "Niederschlag_Uncertainty_Minima",
    "Luftdruck_Value_Maxima",
    "Luftdruck_Value_Minima",
]

# ---------------------------------------------------------------------
# Ergebnisdatei auswählen
# ---------------------------------------------------------------------
# Hier wird festgelegt, welche CSV ausgewertet wird.
# Zum Wechseln zwischen Test-, Pilot- und Finaldaten einfach genau EINE
# der folgenden Zeilen aktiv lassen bzw. den Dateinamen anpassen.

CSV_FILENAME = "Nutzerstudie_all_tidy_Test.csv"       # aktuelle Testdaten
# CSV_FILENAME = "Nutzerstudie_all_tidy_Pilot.csv"  # spätere Pilotstudie
# CSV_FILENAME = "Nutzerstudie_all_tidy_Final.csv"  # spätere Hauptstudie

OUTPUT_FOLDER = "analysis_output"

# Standardmäßig nur vollständig abgeschlossene Teilnehmende analysieren.
INCLUDE_INCOMPLETE = False

# Definition einer "exakten" Antwort
#
# "exact" ist ausschließlich eine DESKRIPTIVE Zusatzkennzahl, z.B.
# "VSUP: 5/20 exakt". Sie geht NICHT in ANOVA, t-Test, Friedman oder
# Wilcoxon ein. Die inferenzstatistischen Tests verwenden weiterhin
# ausschließlich die metrische absolute Abweichung abs_error.
#
# Eine Antwort gilt als exakt, wenn ihre absolute Abweichung vom
# tatsächlichen Extremwert höchstens 1 % der gesamten Wertspanne
# der jeweiligen Variable beträgt.
EXACT_PERCENTAGE = 0.01


# Tatsächliche Werte der Regionen.
# "deviation" = absolute Abweichung vom Zielwert -12 °C.
REGION_VALUES = {
    "VSUP": {
        "R1": {
            "temperature_mean": -13.1901063,
            "deviation_from_target": 1.1901063,
            "uncertainty": 5.4355829,
        },
        "R2": {
            "temperature_mean": -1.23907347408,
            "deviation_from_target": 10.76092652592,
            "uncertainty": 3.135608284,
        },
        "R3": {
            "temperature_mean": 0.46065072160,
            "deviation_from_target": 12.46065072160,
            "uncertainty": 1.4988802804,
        },
        "R4": {
            "temperature_mean": -20.09617172000,
            "deviation_from_target": 8.09617172000,
            "uncertainty": 1.429650956,
        },
        "R5": {
            "temperature_mean": 2.70402026400,
            "deviation_from_target": 14.70402026400,
            "uncertainty": 0.6453833008,
        },
    },
    "ScaledGlyph": {
        "R1": {
            "temperature_mean": -13.324312746,
            "deviation_from_target": 1.324312746,
            "uncertainty": 5.06290306,
        },
        "R2": {
            "temperature_mean": -0.9009721432,
            "deviation_from_target": 11.0990278568,
            "uncertainty": 3.080212368,
        },
        "R3": {
            "temperature_mean": 0.619351364,
            "deviation_from_target": 12.619351364,
            "uncertainty": 1.5227066124,
        },
        "R4": {
            "temperature_mean": -19.99693754000,
            "deviation_from_target": 7.99693754000,
            "uncertainty": 1.455844682,
        },
        "R5": {
            "temperature_mean": 2.77247876000,
            "deviation_from_target": 14.77247876000,
            "uncertainty": 0.650520978,
        },
    },
    "IsoGlyph": {
        "R1": {
            "temperature_mean": -13.1901063,
            "deviation_from_target": 1.1901063,
            "uncertainty": 5.4355829,
        },
        "R2": {
            "temperature_mean": -0.94865091016,
            "deviation_from_target": 11.05134908984,
            "uncertainty": 3.129501516,
        },
        "R3": {
            "temperature_mean": 0.58398272800,
            "deviation_from_target": 12.58398272800,
            "uncertainty": 1.5260805704,
        },
        "R4": {
            "temperature_mean": -3.604896392,
            "deviation_from_target": 8.395103608,
            "uncertainty": 1.424277884,
        },
        "R5": {
            "temperature_mean": 2.78459602000,
            "deviation_from_target": 14.78459602000,
            "uncertainty": 0.6421452136,
        },
    },
}


H1_TRIAL_RE = re.compile(
    r"^(Niederschlag_Uncertainty_(?:Maxima|Minima)|"
    r"Luftdruck_Value_(?:Maxima|Minima))_"
    r"(VSUP|ScaledGlyph|IsoGlyph)$"
)

REGION_TRIALS = {
    "Temperatur_Regions_VSUP": "VSUP",
    "Temperatur_Regions_ScaledGlyph": "ScaledGlyph",
    "Temperatur_Regions_IsoGlyph": "IsoGlyph",
}

REGION_CHOICE_PROMPT = "Welche Region würden Sie wählen?"
REGION_TEXT_PROMPT = "Wieso haben Sie sich für diese Region entschieden?"
CONFIDENCE_PROMPT = "Wie sicher sind Sie sich mit Ihrer Antwort?"
GENERAL_COMMENT_PROMPT = "Weitere Anmerkungen"


# ---------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------

def numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def find_csv(script_dir: Path) -> Path:
    if CSV_FILENAME:
        p = script_dir / CSV_FILENAME
        if not p.exists():
            raise FileNotFoundError(f"CSV nicht gefunden: {p}")
        return p

    candidates = [
        p for p in script_dir.glob("*.csv")
        if not p.name.startswith(("h1_", "h2_", "text_", "missing_"))
        and "analysis_output" not in str(p)
    ]

    # tidy-Dateien bevorzugen
    tidy = [p for p in candidates if "tidy" in p.name.lower()]
    if len(tidy) == 1:
        return tidy[0]

    if len(tidy) > 1:
        raise RuntimeError(
            "Mehrere tidy-CSV-Dateien gefunden:\n"
            + "\n".join(f"  - {p.name}" for p in tidy)
            + "\nBitte oben CSV_FILENAME setzen."
        )

    if len(candidates) == 1:
        return candidates[0]

    if not candidates:
        raise FileNotFoundError(
            "Keine Ergebnis-CSV im selben Ordner wie die .py-Datei gefunden."
        )

    raise RuntimeError(
        "Mehrere CSV-Dateien gefunden:\n"
        + "\n".join(f"  - {p.name}" for p in candidates)
        + "\nBitte oben CSV_FILENAME setzen."
    )


def completed_participants(df: pd.DataFrame) -> set[str]:
    pc = numeric(df["percentComplete"])
    max_pc = df.assign(_pc=pc).groupby("participantId")["_pc"].max()
    return set(max_pc[max_pc >= 100].index.astype(str))


def prepare_base(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["participantId"] = out["participantId"].astype(str)

    if not INCLUDE_INCOMPLETE:
        keep = completed_participants(out)
        out = out[out["participantId"].isin(keep)].copy()

    return out


# ---------------------------------------------------------------------
# Forschungsfrage 1: Extrema / Wahrnehmung
# ---------------------------------------------------------------------

def h1_variable_from_task(task: str) -> str:
    """
    Ordnet die vier Extrema-Aufgaben ihrer jeweiligen Variable zu.
    Maxima und Minima derselben Variable erhalten dadurch dieselbe
    prozentuale Exaktheits-Toleranz.
    """
    if str(task).startswith("Niederschlag_Uncertainty_"):
        return "Niederschlag_Uncertainty"
    if str(task).startswith("Luftdruck_Value_"):
        return "Luftdruck_Value"
    return str(task)


def apply_automatic_exact_tolerance(h1: pd.DataFrame) -> pd.DataFrame:
    """
    Berechnet für jede Variable:
        Toleranz = EXACT_PERCENTAGE * (korrektes Maximum - korrektes Minimum)

    Die korrekten Extrema werden direkt aus correctAnswer der Studie abgeleitet.
    Die Kennzahl 'exact' bleibt rein deskriptiv und beeinflusst keinen Test.
    """
    if h1.empty:
        return h1

    out = h1.copy()
    out["variable"] = out["task"].map(h1_variable_from_task)

    ranges = (
        out.groupby("variable")["correctAnswer"]
        .agg(correct_min="min", correct_max="max")
        .reset_index()
    )
    ranges["value_range"] = ranges["correct_max"] - ranges["correct_min"]
    ranges["exact_tolerance"] = ranges["value_range"] * EXACT_PERCENTAGE

    out = out.drop(columns=["exact_tolerance", "exact"], errors="ignore").merge(
        ranges[
            [
                "variable",
                "correct_min",
                "correct_max",
                "value_range",
                "exact_tolerance",
            ]
        ],
        on="variable",
        how="left",
    )

    # Als pandas Nullable-Boolean anlegen, damit fehlende Werte (pd.NA)
    # auch mit neueren pandas-Versionen sauber gesetzt werden können.
    exact = (out["abs_error"] <= out["exact_tolerance"]).astype("boolean")
    invalid_exact = (
        out["abs_error"].isna()
        | out["exact_tolerance"].isna()
        | (out["value_range"] <= 0)
    )
    exact = exact.mask(invalid_exact, pd.NA)
    out["exact"] = exact

    return out


def exact_tolerance_summary(h1: pd.DataFrame) -> pd.DataFrame:
    """Dokumentiert die automatisch verwendeten 1%-Toleranzen."""
    if h1.empty or "variable" not in h1.columns:
        return pd.DataFrame()

    cols = [
        "variable",
        "correct_min",
        "correct_max",
        "value_range",
        "exact_tolerance",
    ]
    result = h1[cols].drop_duplicates().copy()
    result["exact_percentage"] = EXACT_PERCENTAGE
    return result.sort_values("variable").reset_index(drop=True)


def build_h1(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    cand = df[df["status"].eq("completed")].copy()

    for _, r in cand.iterrows():
        tid = str(r.get("trialId", ""))
        m = H1_TRIAL_RE.match(tid)
        if not m:
            continue

        task, method = m.groups()

        correct = pd.to_numeric(
            pd.Series([r.get("correctAnswer")]),
            errors="coerce"
        ).iloc[0]

        # Nur die eigentliche Klickantwort analysieren.
        # Zeilen wie subjektive Sicherheit besitzen keinen numerischen correctAnswer.
        if pd.isna(correct):
            continue

        answer = pd.to_numeric(
            pd.Series([r.get("answer")]),
            errors="coerce"
        ).iloc[0]

        unanswered = pd.isna(answer) or answer == -999

        if unanswered:
            answer = np.nan
            abs_error = np.nan
        else:
            abs_error = abs(float(answer) - float(correct))

        # Die Toleranz für "exakt" wird erst nach dem Einlesen aller
        # H1-Aufgaben automatisch aus 1 % der Variablenspanne berechnet.
        tolerance = np.nan
        exact = pd.NA

        duration_clean = pd.to_numeric(
            pd.Series([r.get("cleanedDuration")]), errors="coerce"
        ).iloc[0]

        duration_raw = pd.to_numeric(
            pd.Series([r.get("duration")]), errors="coerce"
        ).iloc[0]

        rows.append(
            {
                "participantId": str(r["participantId"]),
                "task": task,
                "visualization": method,
                "answer": answer,
                "correctAnswer": float(correct),
                "abs_error": abs_error,
                "answered": not unanswered,
                "exact_tolerance": tolerance,
                "exact": exact,
                "duration_ms": (
                    duration_clean
                    if pd.notna(duration_clean)
                    else duration_raw
                ),
                "trialId": tid,
                "trialOrder": r.get("trialOrder"),
            }
        )

    h1 = pd.DataFrame(rows)
    if not h1.empty:
        h1 = apply_automatic_exact_tolerance(h1)
    return h1


def h1_complete_triplets(h1: pd.DataFrame) -> pd.DataFrame:
    answered = h1[h1["answered"]].copy()

    counts = (
        answered.groupby(["participantId", "task"])["visualization"]
        .nunique()
        .rename("n_methods")
        .reset_index()
    )

    complete = counts[
        counts["n_methods"] == len(METHOD_ORDER)
    ][["participantId", "task"]]

    return answered.merge(
        complete,
        on=["participantId", "task"],
        how="inner"
    )


def participant_method_aggregate(paired_h1: pd.DataFrame) -> pd.DataFrame:
    if paired_h1.empty:
        return pd.DataFrame()

    agg = (
        paired_h1.groupby(
            ["participantId", "visualization"],
            as_index=False
        )
        .agg(
            mean_abs_error=("abs_error", "mean"),
            median_abs_error=("abs_error", "median"),
            mean_duration_ms=("duration_ms", "mean"),
            n_tasks=("task", "nunique"),
        )
    )

    exact_source = paired_h1.dropna(subset=["exact"])

    if not exact_source.empty:
        ex = (
            exact_source.groupby(
                ["participantId", "visualization"],
                as_index=False
            )
            .agg(exact_rate=("exact", "mean"))
        )
        agg = agg.merge(
            ex,
            on=["participantId", "visualization"],
            how="left"
        )

    return agg


def h1_summary(h1: pd.DataFrame) -> pd.DataFrame:
    if h1.empty:
        return pd.DataFrame()

    grouped = (
        h1.groupby("visualization", as_index=False)
        .agg(
            n_trials=("trialId", "size"),
            n_answered=("answered", "sum"),
            mean_abs_error=("abs_error", "mean"),
            median_abs_error=("abs_error", "median"),
            mean_duration_ms=("duration_ms", "mean"),
            median_duration_ms=("duration_ms", "median"),
        )
    )

    grouped["answer_rate"] = (
        grouped["n_answered"] / grouped["n_trials"]
    )

    if not h1.empty and h1["exact"].notna().any():
        ex_source = h1.dropna(subset=["exact"]).copy()
        ex_source["exact_int"] = ex_source["exact"].astype(int)

        ex = (
            ex_source
            .groupby("visualization", as_index=False)
            .agg(
                n_exact=("exact_int", "sum"),
                n_exact_evaluable=("exact_int", "size"),
                exact_rate=("exact_int", "mean"),
            )
        )
        ex["exact_display"] = (
            ex["n_exact"].astype(str)
            + "/"
            + ex["n_exact_evaluable"].astype(str)
        )

        grouped = grouped.merge(ex, on="visualization", how="left")

    return grouped


# ---------------------------------------------------------------------
# Statistische Tests
# ---------------------------------------------------------------------
#
# Erwarteter Hauptpfad:
# Bei annähernd normalverteilten paarweisen Differenzen:
#   Repeated-Measures-ANOVA -> bei Signifikanz gepaarte t-Tests
#
# Fallback bei verletzter Normalitätsannahme:
#   Friedman-Test -> bei Signifikanz Wilcoxon-Tests
#
# Die Entscheidung wird nicht vorab erzwungen, sondern anhand der Daten
# dokumentiert. So bleibt die Analyse auch für Pilot- und Finaldaten robust.

def normality_of_pairwise_differences(
    wide: pd.DataFrame,
    methods: list[str]
) -> pd.DataFrame:
    rows = []

    for i in range(len(methods)):
        for j in range(i + 1, len(methods)):
            a, b = methods[i], methods[j]
            pair = wide[[a, b]].dropna()
            diff = pair[a] - pair[b]

            if len(diff) >= 3:
                w, p = stats.shapiro(diff)
            else:
                w, p = np.nan, np.nan

            rows.append(
                {
                    "comparison": f"{a} - {b}",
                    "n": len(diff),
                    "shapiro_W": w,
                    "shapiro_p": p,
                    "normal_at_0.05": (
                        bool(p >= 0.05)
                        if pd.notna(p)
                        else pd.NA
                    ),
                }
            )

    return pd.DataFrame(rows)


def omnibus_and_posthoc(
    long_df: pd.DataFrame,
    value_col: str,
    subject_col: str = "participantId",
    method_col: str = "visualization",
):
    x = long_df[
        [subject_col, method_col, value_col]
    ].dropna().copy()

    wide = x.pivot_table(
        index=subject_col,
        columns=method_col,
        values=value_col,
        aggfunc="mean",
    )

    methods = [m for m in METHOD_ORDER if m in wide.columns]
    wide = wide[methods].dropna()

    if len(methods) != 3 or len(wide) < 3:
        note = pd.DataFrame(
            [{
                "test": "nicht berechnet",
                "n_complete_participants": len(wide),
                "statistic": np.nan,
                "df": np.nan,
                "p": np.nan,
                "significant_0.05": pd.NA,
                "note": (
                    "Zu wenige vollständige Fälle oder "
                    "nicht alle drei Methoden vorhanden."
                ),
            }]
        )
        return pd.DataFrame(), note, pd.DataFrame()

    normality = normality_of_pairwise_differences(wide, methods)
    all_normal = normality["normal_at_0.05"].fillna(False).all()

    if all_normal:
        long_complete = (
            wide.reset_index()
            .melt(
                id_vars=subject_col,
                var_name=method_col,
                value_name=value_col,
            )
        )

        model = AnovaRM(
            long_complete,
            depvar=value_col,
            subject=subject_col,
            within=[method_col],
        ).fit()

        table = model.anova_table.reset_index()

        omnibus = pd.DataFrame(
            [{
                "test": "Repeated Measures ANOVA",
                "n_complete_participants": len(wide),
                "statistic": float(table.loc[0, "F Value"]),
                "df": (
                    f"{float(table.loc[0, 'Num DF']):.0f}, "
                    f"{float(table.loc[0, 'Den DF']):.0f}"
                ),
                "p": float(table.loc[0, "Pr > F"]),
                "significant_0.05": bool(
                    table.loc[0, "Pr > F"] < 0.05
                ),
                "note": (
                    "Verwendet, weil alle paarweisen Differenzen "
                    "Shapiro p >= .05 hatten."
                ),
            }]
        )
        posthoc_type = "paired t-test"

    else:
        vals = [wide[m].to_numpy() for m in methods]
        stat, p = stats.friedmanchisquare(*vals)

        omnibus = pd.DataFrame(
            [{
                "test": "Friedman-Test",
                "n_complete_participants": len(wide),
                "statistic": float(stat),
                "df": len(methods) - 1,
                "p": float(p),
                "significant_0.05": bool(p < 0.05),
                "note": (
                    "Verwendet, weil mindestens eine paarweise "
                    "Differenz Shapiro p < .05 hatte."
                ),
            }]
        )
        posthoc_type = "Wilcoxon signed-rank"

    if not bool(omnibus.loc[0, "significant_0.05"]):
        posthoc = pd.DataFrame(
            [{
                "test": posthoc_type,
                "comparison": "nicht durchgeführt",
                "n": len(wide),
                "statistic": np.nan,
                "p_raw": np.nan,
                "p_bonferroni": np.nan,
                "significant_bonferroni_0.05": pd.NA,
                "note": "Gesamtvergleich nicht signifikant.",
            }]
        )
        return normality, omnibus, posthoc

    rows = []
    n_comp = math.comb(len(methods), 2)

    for i in range(len(methods)):
        for j in range(i + 1, len(methods)):
            a, b = methods[i], methods[j]
            xa, xb = wide[a], wide[b]

            if posthoc_type == "paired t-test":
                stat, p = stats.ttest_rel(
                    xa,
                    xb,
                    nan_policy="omit"
                )
            else:
                d = xa - xb
                if np.allclose(d, 0):
                    stat, p = 0.0, 1.0
                else:
                    stat, p = stats.wilcoxon(
                        xa,
                        xb,
                        zero_method="wilcox"
                    )

            p_bonf = min(float(p) * n_comp, 1.0)

            rows.append(
                {
                    "test": posthoc_type,
                    "comparison": f"{a} vs {b}",
                    "n": len(xa),
                    "statistic": float(stat),
                    "p_raw": float(p),
                    "p_bonferroni": p_bonf,
                    "significant_bonferroni_0.05": (
                        p_bonf < 0.05
                    ),
                    "note": (
                        f"Bonferroni über {n_comp} Vergleiche; "
                        f"äquivalent alpha={0.05/n_comp:.4f}."
                    ),
                }
            )

    return normality, omnibus, pd.DataFrame(rows)


# ---------------------------------------------------------------------
# Forschungsfrage 2: Regionsentscheidung
# ---------------------------------------------------------------------

def extract_region_data(df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for pid, psub in df.groupby("participantId"):
        for tid, method in REGION_TRIALS.items():
            sub = psub[
                (psub["trialId"] == tid)
                & (psub["status"] == "completed")
            ]

            if sub.empty:
                continue

            def one_answer(prompt: str):
                s = sub.loc[
                    sub["responsePrompt"] == prompt,
                    "answer"
                ].dropna()
                return s.iloc[0] if len(s) else np.nan

            choice = one_answer(REGION_CHOICE_PROMPT)
            justification = one_answer(REGION_TEXT_PROMPT)
            confidence = one_answer(CONFIDENCE_PROMPT)

            dur = numeric(sub["cleanedDuration"]).dropna()
            if dur.empty:
                dur = numeric(sub["duration"]).dropna()

            rows.append(
                {
                    "participantId": str(pid),
                    "visualization": method,
                    "region": choice,
                    "justification": justification,
                    "confidence": pd.to_numeric(
                        pd.Series([confidence]),
                        errors="coerce"
                    ).iloc[0],
                    "duration_ms": (
                        dur.iloc[0]
                        if len(dur)
                        else np.nan
                    ),
                    "trialId": tid,
                }
            )

    return pd.DataFrame(rows)


def enrich_regions(region: pd.DataFrame) -> pd.DataFrame:
    x = region.copy()

    means = []
    deviations = []
    uncertainties = []
    safest = []

    for _, r in x.iterrows():
        method = r["visualization"]
        region_name = r["region"]

        values = REGION_VALUES.get(method, {}).get(
            str(region_name),
            None
        )

        if values is None:
            means.append(np.nan)
            deviations.append(np.nan)
            uncertainties.append(np.nan)
            safest.append(pd.NA)
            continue

        mean_v = values["temperature_mean"]
        dev_v = values["deviation_from_target"]
        unc_v = values["uncertainty"]

        method_uncertainties = {
            rr: vv["uncertainty"]
            for rr, vv in REGION_VALUES[method].items()
        }
        min_unc = min(method_uncertainties.values())

        means.append(mean_v)
        deviations.append(dev_v)
        uncertainties.append(unc_v)
        safest.append(bool(np.isclose(unc_v, min_unc)))

    x["chosen_temperature_mean"] = means
    x["deviation_from_target"] = deviations
    x["chosen_uncertainty"] = uncertainties
    x["is_safest_region"] = pd.Series(
        safest,
        dtype="boolean"
    )

    return x


def region_summaries(region: pd.DataFrame):
    freq = (
        region.groupby(
            ["visualization", "region"],
            dropna=False
        )
        .size()
        .rename("n")
        .reset_index()
    )

    freq["proportion_within_method"] = (
        freq.groupby("visualization")["n"]
        .transform(lambda s: s / s.sum())
    )

    summary = (
        region.groupby("visualization", as_index=False)
        .agg(
            n_choices=("region", "count"),
            mean_chosen_uncertainty=(
                "chosen_uncertainty",
                "mean"
            ),
            median_chosen_uncertainty=(
                "chosen_uncertainty",
                "median"
            ),
            mean_deviation_from_target=(
                "deviation_from_target",
                "mean"
            ),
            median_deviation_from_target=(
                "deviation_from_target",
                "median"
            ),
            mean_confidence=("confidence", "mean"),
            median_confidence=("confidence", "median"),
            mean_duration_ms=("duration_ms", "mean"),
            median_duration_ms=("duration_ms", "median"),
        )
    )

    safest = (
        region.dropna(subset=["is_safest_region"])
        .groupby("visualization", as_index=False)
        .agg(
            safest_region_rate=("is_safest_region", "mean"),
            n_safest_evaluable=("is_safest_region", "size"),
        )
    )

    summary = summary.merge(
        safest,
        on="visualization",
        how="left"
    )

    return freq, summary


# ---------------------------------------------------------------------
# Freitext
# ---------------------------------------------------------------------

def build_feedback_output(df: pd.DataFrame) -> pd.DataFrame:
    """
    Enthält ausschließlich das freie Textfeld aus dem Abschluss-Feedback.
    Die Begründungen der H2-Regionsaufgabe bleiben nur in H2 Regionsauswahl
    und werden hier nicht doppelt ausgegeben.
    """
    comments = df[
        (df["trialId"] == "Feedback")
        & (df["status"] == "completed")
        & (df["responsePrompt"] == GENERAL_COMMENT_PROMPT)
    ].copy()

    rows = []
    for _, r in comments.iterrows():
        text = r.get("answer")
        if pd.notna(text) and str(text).strip():
            rows.append(
                {
                    "participantId": str(r["participantId"]),
                    "feedback_prompt": str(r.get("responsePrompt", GENERAL_COMMENT_PROMPT)),
                    "text": str(text).strip(),
                }
            )

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values("participantId").reset_index(drop=True)

    return out


def write_feedback_markdown(
    feedback_df: pd.DataFrame,
    path: Path
) -> None:
    with path.open("w", encoding="utf-8") as f:
        f.write("# Abschlussfeedback\n\n")

        if feedback_df.empty:
            f.write("Keine Freitextantworten im Abschlussfeedback gefunden.\n")
            return

        for _, r in feedback_df.iterrows():
            f.write(f"**{r['participantId']}**\n\n")
            f.write(str(r["text"]).strip() + "\n\n")
            f.write("---\n\n")


def write_missing_report(
    h1: pd.DataFrame,
    region: pd.DataFrame,
    path: Path
) -> None:
    miss_h1 = h1[
        ~h1["answered"]
    ][
        [
            "participantId",
            "task",
            "visualization",
            "trialId"
        ]
    ].copy()

    miss_h1["type"] = (
        "Extrema-Klick nicht beantwortet (-999/NaN)"
    )

    miss_region = region[
        region["region"].isna()
    ][
        [
            "participantId",
            "visualization",
            "trialId"
        ]
    ].copy()

    miss_region["task"] = (
        "Temperatur – Standortentscheidung"
    )
    miss_region["type"] = (
        "Region nicht ausgewählt"
    )

    cols = [
        "participantId",
        "task",
        "visualization",
        "trialId",
        "type",
    ]

    all_missing = pd.concat(
        [
            miss_h1[cols],
            miss_region[cols]
        ],
        ignore_index=True
    )

    export_csv(all_missing, path)


# ---------------------------------------------------------------------
# Zusammenfassung der Testentscheidung
# ---------------------------------------------------------------------

def describe_test_decision(
    section_title: str,
    normality: pd.DataFrame,
    omnibus: pd.DataFrame,
    posthoc: pd.DataFrame,
) -> str:
    lines = [section_title]

    if omnibus.empty:
        lines.append("Keine Testentscheidung berechnet.")
        return "\n".join(lines)

    test_name = str(omnibus.iloc[0].get("test", "unbekannt"))
    lines.append(f"Verwendeter Gesamtvergleich: {test_name}")

    if not normality.empty and "normal_at_0.05" in normality.columns:
        vals = normality["normal_at_0.05"].dropna()
        if len(vals):
            all_normal = bool(vals.all())
            lines.append(
                "Normalitätsannahme der paarweisen Differenzen erfüllt: "
                + ("Ja" if all_normal else "Nein")
            )

        lines.append("Shapiro-Wilk-Prüfungen der paarweisen Differenzen:")
        for _, r in normality.iterrows():
            p = r.get("shapiro_p", np.nan)
            p_text = f"{float(p):.6g}" if pd.notna(p) else "nicht berechnet"
            lines.append(
                f"  {r.get('comparison', '')}: p = {p_text}"
            )

    if "p" in omnibus.columns and pd.notna(omnibus.iloc[0].get("p")):
        p_omni = float(omnibus.iloc[0]["p"])
        lines.append(f"Omnibus-p-Wert: {p_omni:.6g}")
        lines.append(
            "Omnibustest signifikant (alpha=0,05): "
            + ("Ja" if p_omni < 0.05 else "Nein")
        )

    if "Repeated Measures ANOVA" in test_name:
        lines.append("Vorgesehener Post-hoc-Test bei Signifikanz: gepaarter t-Test")
    elif "Friedman" in test_name:
        lines.append("Vorgesehener Post-hoc-Test bei Signifikanz: Wilcoxon signed-rank")

    lines.append("Bonferroni-korrigiertes Alpha bei 3 Paarvergleichen: 0,0167")

    if not posthoc.empty:
        first_comp = str(posthoc.iloc[0].get("comparison", ""))
        if first_comp == "nicht durchgeführt":
            lines.append("Post-hoc-Tests: nicht durchgeführt, da Gesamtvergleich nicht signifikant.")
        elif first_comp:
            lines.append("Post-hoc-Tests: durchgeführt.")

    return "\n".join(lines)


def write_analysis_decision(
    path: Path,
    h1_result,
    h1_duration_result,
    h2_uncertainty_result,
    h2_deviation_result,
) -> None:
    blocks = [
        describe_test_decision(
            "H1 – Genauigkeit / absolute Abweichung",
            *h1_result
        ),
        describe_test_decision(
            "H1 – Antwortzeit (sekundär)",
            *h1_duration_result
        ),
        describe_test_decision(
            "H2 – Unsicherheit der gewählten Region",
            *h2_uncertainty_result
        ),
        describe_test_decision(
            "H2 – Abweichung vom Zielwert -12 °C (zusätzlich)",
            *h2_deviation_result
        ),
    ]

    header = (
        "TESTENTSCHEIDUNGEN DER AUSWERTUNG\n"
        "================================\n\n"
        "Die Auswahl zwischen parametrischem und nichtparametrischem Pfad "
        "erfolgt anhand der Normalität der paarweisen Differenzen.\n"
        "Parametrischer Pfad: Repeated-Measures-ANOVA -> gepaarte t-Tests.\n"
        "Fallback: Friedman-Test -> Wilcoxon-Tests.\n\n"
    )

    path.write_text(
        header + "\n\n".join(blocks) + "\n",
        encoding="utf-8"
    )


# ---------------------------------------------------------------------
# Ausgabe: CSV + XLSX
# ---------------------------------------------------------------------

def export_csv(df: pd.DataFrame, path: Path) -> None:
    """
    CSV für deutsches Excel:
    - Semikolon als Trennzeichen
    - UTF-8 mit BOM, damit Umlaute und Gedankenstriche korrekt erkannt werden
    """
    df.to_csv(
        path,
        index=False,
        sep=";",
        encoding="utf-8-sig",
    )


GERMAN_COLUMN_NAMES = {
    "participantId": "Teilnehmer-ID",
    "task": "Aufgabe",
    "visualization": "Visualisierung",
    "answer": "Antwort",
    "correctAnswer": "Korrekte Antwort",
    "abs_error": "Absolute Abweichung",
    "answered": "Beantwortet",
    "exact_tolerance": "Exaktheits-Toleranz",
    "exact": "Exakt",
    "duration_ms": "Antwortzeit (ms)",
    "duration_raw_ms": "Roh-Antwortzeit (ms)",
    "trialId": "Trial-ID",
    "trialOrder": "Ursprüngliche Reihenfolge",
    "variable": "Variable",
    "correct_min": "Korrektes Minimum",
    "correct_max": "Korrektes Maximum",
    "value_range": "Wertspanne",
    "exact_percentage": "Exakt-Toleranz (Anteil)",
    "n_trials": "Anzahl Aufgaben",
    "n_answered": "Anzahl beantwortet",
    "mean_abs_error": "Mittlere absolute Abweichung",
    "median_abs_error": "Mediane absolute Abweichung",
    "mean_duration_ms": "Mittlere Antwortzeit (ms)",
    "median_duration_ms": "Mediane Antwortzeit (ms)",
    "answer_rate": "Beantwortungsquote",
    "n_exact": "Anzahl exakt",
    "n_exact_evaluable": "Auswertbare Antworten für Exakt",
    "n_evaluable": "Auswertbare Antworten",
    "exact_rate": "Anteil exakt",
    "exact_display": "Exakt (Anzahl)",
    "n_tasks": "Anzahl Aufgaben",
    "comparison": "Vergleich",
    "shapiro_W": "Shapiro-Wilk W",
    "shapiro_p": "Shapiro-Wilk p",
    "normal_at_0.05": "Normalverteilt (α = 0,05)",
    "test": "Test",
    "n_complete_participants": "Vollständige Teilnehmer",
    "statistic": "Teststatistik",
    "df": "Freiheitsgrade",
    "p": "p-Wert",
    "significant_0.05": "Signifikant (α = 0,05)",
    "note": "Hinweis",
    "p_raw": "p-Wert (unkorrigiert)",
    "p_bonferroni": "p-Wert (Bonferroni)",
    "significant_bonferroni_0.05": "Signifikant nach Bonferroni",
    "region": "Gewählte Region",
    "justification": "Begründung",
    "confidence": "Subjektive Sicherheit",
    "chosen_temperature_mean": "Gewählter Temperaturmittelwert (°C)",
    "deviation_from_target": "Abweichung vom Zielwert -12 °C",
    "chosen_uncertainty": "Unsicherheit der gewählten Region",
    "is_safest_region": "Sicherste Region gewählt",
    "n_choices": "Anzahl Entscheidungen",
    "mean_chosen_uncertainty": "Mittlere gewählte Unsicherheit",
    "median_chosen_uncertainty": "Mediane gewählte Unsicherheit",
    "mean_deviation_from_target": "Mittlere Zielabweichung",
    "median_deviation_from_target": "Mediane Zielabweichung",
    "mean_confidence": "Mittlere subjektive Sicherheit",
    "median_confidence": "Mediane subjektive Sicherheit",
    "safest_region_rate": "Anteil sicherste Region gewählt",
    "n_safest_evaluable": "Auswertbare Entscheidungen",
    "n": "Anzahl",
    "proportion_within_method": "Anteil innerhalb Methode",
    "feedback_prompt": "Feedback-Frage",
    "text": "Feedback-Text",
    "type": "Typ",
}


def germanize_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    return df.copy().rename(columns=GERMAN_COLUMN_NAMES)


def sort_h1_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Feste, lesbare Reihenfolge je Teilnehmer; trialOrder bleibt sichtbar."""
    if df is None or df.empty:
        return df.copy() if df is not None else pd.DataFrame()

    out = df.copy()
    task_order = {name: i for i, name in enumerate(H1_TASK_ORDER)}
    method_order = {name: i for i, name in enumerate(METHOD_ORDER)}

    out["_task_order"] = out["task"].map(task_order).fillna(999)
    out["_method_order"] = out["visualization"].map(method_order).fillna(999)

    out = (
        out.sort_values(
            ["participantId", "_task_order", "_method_order", "trialOrder"],
            na_position="last"
        )
        .drop(columns=["_task_order", "_method_order"])
        .reset_index(drop=True)
    )
    return out


def sort_h2_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df.copy() if df is not None else pd.DataFrame()

    out = df.copy()
    method_order = {name: i for i, name in enumerate(METHOD_ORDER)}
    out["_method_order"] = out["visualization"].map(method_order).fillna(999)
    out = (
        out.sort_values(["participantId", "_method_order"])
        .drop(columns="_method_order")
        .reset_index(drop=True)
    )
    return out


def add_blank_rows_between_groups(
    df: pd.DataFrame,
    group_col: str
) -> pd.DataFrame:
    """Nur für die Excel-Darstellung: Leerzeile zwischen Teilnehmern."""
    if df is None or df.empty or group_col not in df.columns:
        return df.copy() if df is not None else pd.DataFrame()

    parts = []
    groups = list(df.groupby(group_col, sort=False, dropna=False))

    for i, (_, group) in enumerate(groups):
        parts.append(group)
        if i < len(groups) - 1:
            parts.append(pd.DataFrame([{col: None for col in df.columns}]))

    return pd.concat(parts, ignore_index=True)


def excel_sheet(
    df: pd.DataFrame,
    description: str | None = None,
    blank_between: str | None = None,
) -> dict:
    return {
        "description": description,
        "df": df,
        "blank_between": blank_between,
    }


def excel_sections(
    sections: list[tuple[str, pd.DataFrame, str | None]],
    description: str | None = None,
) -> dict:
    return {
        "description": description,
        "sections": sections,
    }


def _xlsx_column_width(df: pd.DataFrame, col: str) -> int:
    values = df[col].astype("string").fillna("") if col in df.columns else pd.Series([], dtype="string")
    max_len = max([len(str(col))] + [len(v) for v in values.head(500)])
    is_text = col in {
        "Hinweis", "Begründung", "Feedback-Text", "Typ",
        "Trial-ID", "Aufgabe"
    }
    return min(max(max_len + 5, 15), 70 if is_text else 38)


def export_results_xlsx(
    path: Path,
    sheets: dict[str, dict],
) -> None:
    """
    Zentrale, lesbare Excel-Datei.
    Funktioniert sowohl mit xlsxwriter als auch mit openpyxl.

    Layout:
    - keine unnötigen Leerzeilen am Blattanfang
    - Beschreibung (falls vorhanden) direkt in Zeile 1
    - Tabellenüberschrift direkt darunter
    - großzügige Spaltenbreiten inklusive Platz für Filterpfeile
    - deutsche Spaltennamen
    - zusammengehörige Statistikschritte auf einem Blatt
    """
    try:
        writer = pd.ExcelWriter(path, engine="xlsxwriter")
    except ModuleNotFoundError:
        writer = pd.ExcelWriter(path, engine="openpyxl")

    with writer:
        for sheet_name, spec in sheets.items():
            safe_name = sheet_name[:31]
            description = spec.get("description")

            # ==========================================================
            # XLSXWRITER
            # ==========================================================
            if writer.engine == "xlsxwriter":
                workbook = writer.book
                worksheet = workbook.add_worksheet(safe_name)
                writer.sheets[safe_name] = worksheet

                description_format = workbook.add_format({
                    "italic": True,
                    "text_wrap": True,
                    "valign": "top",
                })
                section_format = workbook.add_format({
                    "bold": True,
                    "font_size": 11,
                    "bottom": 1,
                })
                header_format = workbook.add_format({
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "border": 1,
                })
                text_format = workbook.add_format({
                    "text_wrap": True,
                    "valign": "top",
                })
                percent_format = workbook.add_format({
                    "num_format": "0.00%",
                })

                current_row = 0
                used_frames = []

                # Beschreibung direkt in die erste Zeile; KEINE Leerzeile davor.
                if description:
                    worksheet.write(current_row, 0, description, description_format)
                    current_row += 1

                if "sections" in spec:
                    for section_index, (title, frame, section_desc) in enumerate(spec["sections"]):
                        frame = germanize_columns(frame)
                        used_frames.append(frame)

                        # Nur zwischen Sektionen eine Leerzeile.
                        if section_index > 0:
                            current_row += 1

                        worksheet.write(current_row, 0, title, section_format)
                        current_row += 1

                        if section_desc:
                            worksheet.write(current_row, 0, section_desc, description_format)
                            current_row += 1

                        header_row = current_row

                        for col_num, col in enumerate(frame.columns):
                            worksheet.write(header_row, col_num, col, header_format)

                        if not frame.empty:
                            frame.to_excel(
                                writer,
                                sheet_name=safe_name,
                                index=False,
                                header=False,
                                startrow=header_row + 1,
                            )

                        current_row = header_row + len(frame) + 1

                else:
                    frame = spec.get("df", pd.DataFrame()).copy()

                    if spec.get("blank_between"):
                        frame = add_blank_rows_between_groups(
                            frame,
                            spec["blank_between"]
                        )

                    frame = germanize_columns(frame)
                    used_frames.append(frame)

                    header_row = current_row

                    for col_num, col in enumerate(frame.columns):
                        worksheet.write(header_row, col_num, col, header_format)

                    if not frame.empty:
                        frame.to_excel(
                            writer,
                            sheet_name=safe_name,
                            index=False,
                            header=False,
                            startrow=header_row + 1,
                        )

                    if len(frame.columns) > 0:
                        worksheet.freeze_panes(header_row + 1, 0)
                        worksheet.autofilter(
                            header_row,
                            0,
                            header_row + max(len(frame), 1),
                            len(frame.columns) - 1,
                        )

                # Spaltenbreite: Überschrift + bis zu 500 Werte berücksichtigen.
                all_columns = []
                for frame in used_frames:
                    for col in frame.columns:
                        if col not in all_columns:
                            all_columns.append(col)

                for col_num, col in enumerate(all_columns):
                    max_len = len(str(col))

                    for frame in used_frames:
                        if col not in frame.columns:
                            continue
                        values = frame[col].astype("string").fillna("")
                        if not values.empty:
                            max_len = max(
                                max_len,
                                max(len(str(v)) for v in values.head(500))
                            )

                    is_text = col in {
                        "Hinweis", "Begründung", "Feedback-Text",
                        "Typ", "Aufgabe", "Trial-ID"
                    }

                    # Deutlich mehr Reserve für Dropdown/Filter-Symbol.
                    width = min(
                        max(max_len + 8, 18),
                        75 if is_text else 45
                    )

                    fmt = text_format if is_text else None
                    if "Anteil" in col or "Quote" in col:
                        fmt = percent_format

                    worksheet.set_column(col_num, col_num, width, fmt)

                # Kopfzeilen etwas höher, damit Umbruch nicht abgeschnitten wird.
                worksheet.set_default_row(18)

            # ==========================================================
            # OPENPYXL-FALLBACK
            # ==========================================================
            else:
                from openpyxl.styles import Font, Alignment
                from openpyxl.utils import get_column_letter

                current_row = 0
                written_frames = []

                # Beschreibung wirklich schreiben statt nur zwei Zeilen freizuhalten.
                if description:
                    pd.DataFrame([[description]]).to_excel(
                        writer,
                        sheet_name=safe_name,
                        index=False,
                        header=False,
                        startrow=current_row,
                    )
                    current_row += 1

                if "sections" in spec:
                    for section_index, (title, frame, section_desc) in enumerate(spec["sections"]):
                        frame = germanize_columns(frame)
                        written_frames.append(frame)

                        if section_index > 0:
                            current_row += 1

                        pd.DataFrame([[title]]).to_excel(
                            writer,
                            sheet_name=safe_name,
                            index=False,
                            header=False,
                            startrow=current_row,
                        )
                        current_row += 1

                        if section_desc:
                            pd.DataFrame([[section_desc]]).to_excel(
                                writer,
                                sheet_name=safe_name,
                                index=False,
                                header=False,
                                startrow=current_row,
                            )
                            current_row += 1

                        frame.to_excel(
                            writer,
                            sheet_name=safe_name,
                            index=False,
                            startrow=current_row,
                        )
                        current_row += len(frame) + 2

                else:
                    frame = spec.get("df", pd.DataFrame()).copy()

                    if spec.get("blank_between"):
                        frame = add_blank_rows_between_groups(
                            frame,
                            spec["blank_between"]
                        )

                    frame = germanize_columns(frame)
                    written_frames.append(frame)

                    frame.to_excel(
                        writer,
                        sheet_name=safe_name,
                        index=False,
                        startrow=current_row,
                    )

                worksheet = writer.sheets[safe_name]

                # Lesbare Überschriften / Beschreibungen.
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        cell.alignment = Alignment(
                            vertical="top",
                            wrap_text=True
                        )

                # Erste nichtleere Tabellenkopfzeilen fett formatieren.
                # Alle Zellen mit bekannten deutschen Spaltennamen erkennen.
                known_headers = set(GERMAN_COLUMN_NAMES.values())
                for row in worksheet.iter_rows():
                    header_hits = sum(
                        1 for cell in row if cell.value in known_headers
                    )
                    if header_hits >= 1:
                        for cell in row:
                            if cell.value is not None:
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(
                                    vertical="top",
                                    wrap_text=True
                                )

                # Spaltenbreite mit großzügigem Aufschlag für Filterpfeile.
                max_col = worksheet.max_column
                for col_idx in range(1, max_col + 1):
                    max_len = 0

                    for row_idx in range(1, min(worksheet.max_row, 500) + 1):
                        value = worksheet.cell(row_idx, col_idx).value
                        if value is not None:
                            max_len = max(max_len, len(str(value)))

                    width = min(max(max_len + 8, 18), 75)
                    worksheet.column_dimensions[
                        get_column_letter(col_idx)
                    ].width = width

                # Nur bei einfachen Tabellen filtern/fixieren.
                if "sections" not in spec:
                    header_row_excel = 2 if description else 1
                    if worksheet.max_column > 0:
                        worksheet.freeze_panes = f"A{header_row_excel + 1}"
                        worksheet.auto_filter.ref = (
                            f"A{header_row_excel}:"
                            f"{get_column_letter(worksheet.max_column)}"
                            f"{worksheet.max_row}"
                        )



# ---------------------------------------------------------------------
# Hauptprogramm
# ---------------------------------------------------------------------

def main():
    script_dir = Path(__file__).resolve().parent
    csv_path = find_csv(script_dir)

    output_dir = script_dir / OUTPUT_FOLDER
    output_dir.mkdir(exist_ok=True)

    print(f"CSV: {csv_path.name}")
    print(f"Ausgabeordner: {output_dir}\n")

    df = pd.read_csv(csv_path)
    base = prepare_base(df)

    # H1 ---------------------------------------------------------------
    h1 = build_h1(base)
    export_csv(
        h1,
        output_dir / "h1_extrema_trial_level.csv"
    )

    exact_tol = exact_tolerance_summary(h1)
    export_csv(
        exact_tol,
        output_dir / "h1_exact_tolerances.csv"
    )

    h1_sum = h1_summary(h1)
    export_csv(
        h1_sum,
        output_dir / "h1_summary_by_visualization.csv"
    )

    # Rein deskriptive Übersicht der exakten Antworten.
    # Diese Datei beeinflusst KEINEN inferenzstatistischen Test.
    exact_by_task = pd.DataFrame()
    if h1["exact"].notna().any():
        exact_source = h1.dropna(subset=["exact"]).copy()
        exact_source["exact_int"] = exact_source["exact"].astype(int)

        exact_by_task = (
            exact_source
            .groupby(["task", "visualization"], as_index=False)
            .agg(
                n_exact=("exact_int", "sum"),
                n_evaluable=("exact_int", "size"),
            )
        )
        exact_by_task["exact_rate"] = (
            exact_by_task["n_exact"] / exact_by_task["n_evaluable"]
        )
        exact_by_task["exact_display"] = (
            exact_by_task["n_exact"].astype(str)
            + "/"
            + exact_by_task["n_evaluable"].astype(str)
        )

        export_csv(
            exact_by_task,
            output_dir / "h1_exact_answers_descriptive.csv"
        )

    paired_h1 = h1_complete_triplets(h1)
    export_csv(
        paired_h1,
        output_dir / "h1_paired_task_level.csv"
    )

    pma = participant_method_aggregate(paired_h1)
    export_csv(
        pma,
        output_dir / "h1_participant_method.csv"
    )

    normality = pd.DataFrame()
    omnibus = pd.DataFrame()
    posthoc = pd.DataFrame()
    n_t = pd.DataFrame()
    o_t = pd.DataFrame()
    p_t = pd.DataFrame()

    h1_result = (normality, omnibus, posthoc)
    h1_duration_result = (n_t, o_t, p_t)

    if not pma.empty:
        normality, omnibus, posthoc = (
            omnibus_and_posthoc(
                pma,
                "mean_abs_error"
            )
        )
        h1_result = (normality, omnibus, posthoc)

        export_csv(
            normality,
            output_dir / "h1_normality_pairwise_differences.csv"
        )
        export_csv(
            omnibus,
            output_dir / "h1_omnibus_test.csv"
        )
        export_csv(
            posthoc,
            output_dir / "h1_posthoc_tests.csv"
        )

        # Sekundäre Leistungskennzahl Antwortzeit
        n_t, o_t, p_t = omnibus_and_posthoc(
            pma,
            "mean_duration_ms"
        )
        h1_duration_result = (n_t, o_t, p_t)

        export_csv(
            n_t,
            output_dir / "h1_duration_normality.csv"
        )
        export_csv(
            o_t,
            output_dir / "h1_duration_omnibus_test.csv"
        )
        export_csv(
            p_t,
            output_dir / "h1_duration_posthoc_tests.csv"
        )

    # H2 ---------------------------------------------------------------
    region = extract_region_data(base)
    region = enrich_regions(region)

    export_csv(
        region,
        output_dir / "h2_region_choices_long.csv"
    )

    freq, summary = region_summaries(region)

    export_csv(
        freq,
        output_dir / "h2_region_choice_frequencies.csv"
    )

    export_csv(
        summary,
        output_dir / "h2_summary_by_visualization.csv"
    )

    # Hauptmessgröße H2:
    # Unsicherheit der tatsächlich gewählten Region.
    # Es wird exakt derselbe Testpfad wie bei H1 verwendet:
    # Normalitätsprüfung -> Repeated-Measures-ANOVA + gepaarte t-Tests
    # oder bei verletzter Normalitätsannahme Friedman + Wilcoxon.
    n_r, o_r, p_r = omnibus_and_posthoc(
        region,
        "chosen_uncertainty"
    )
    h2_uncertainty_result = (n_r, o_r, p_r)

    export_csv(
        n_r,
        output_dir / "h2_uncertainty_normality.csv"
    )
    export_csv(
        o_r,
        output_dir / "h2_uncertainty_omnibus_test.csv"
    )
    export_csv(
        p_r,
        output_dir / "h2_uncertainty_posthoc_tests.csv"
    )

    # Zusätzlich: Abweichung vom gewünschten Temperaturwert -12 °C
    n_d, o_d, p_d = omnibus_and_posthoc(
        region,
        "deviation_from_target"
    )
    h2_deviation_result = (n_d, o_d, p_d)

    export_csv(
        n_d,
        output_dir / "h2_deviation_normality.csv"
    )
    export_csv(
        o_d,
        output_dir / "h2_deviation_omnibus_test.csv"
    )
    export_csv(
        p_d,
        output_dir / "h2_deviation_posthoc_tests.csv"
    )

    # Abschlussfeedback -------------------------------------------------
    feedback_df = build_feedback_output(base)

    export_csv(
        feedback_df,
        output_dir / "feedback.csv"
    )

    write_feedback_markdown(
        feedback_df,
        output_dir / "feedback.md"
    )

    # Alte, inzwischen überflüssige Freitextdateien aus früheren Versionen
    # entfernen, damit im Ausgabeordner keine doppelten H2-Begründungen bleiben.
    for obsolete_name in [
        "text_answers_by_task.csv",
        "text_answers_by_task.md",
    ]:
        obsolete_path = output_dir / obsolete_name
        if obsolete_path.exists():
            obsolete_path.unlink()

    write_missing_report(
        h1,
        region,
        output_dir /
        "missing_answers.csv"
    )

    write_analysis_decision(
        output_dir / "analysis_decision.txt",
        h1_result,
        h1_duration_result,
        h2_uncertainty_result,
        h2_deviation_result,
    )

    # Zentrale Excel-Arbeitsmappe mit allen Tabellen.
    # Leere Tabellenblätter sind bei den Testdaten erwartbar, wenn z.B.
    # für H1 keine gültigen Antworten vorhanden sind.
    # Excel-spezifische Sortierung: feste Reihenfolge für bessere Lesbarkeit.
    h1_excel = sort_h1_for_excel(h1)
    paired_h1_excel = sort_h1_for_excel(paired_h1)
    region_excel = sort_h2_for_excel(region)

    workbook_sheets = {
        "H1 Übersicht": excel_sheet(
            h1_sum,
            description=(
                "Deskriptive Zusammenfassung aller H1-Einzelantworten je Visualisierung. "
                "Hier werden auch Antworten aus unvollständigen Dreiergruppen berücksichtigt, "
                "solange die jeweilige Einzelaufgabe beantwortet wurde. "
                "Die inferenzstatistischen Methodenvergleiche basieren dagegen ausschließlich "
                "auf den vollständig gepaarten Aufgaben aus 'H1 Gepaarte Antworten'. "
                "'Anzahl' bezeichnet jeweils die Zahl der berücksichtigten Beobachtungen."
            ),
        ),

        # Direkt nebeneinander: vollständige Einzelantworten und daraus
        # gefilterte gepaarte Antworten.
        "H1 Einzelantworten": excel_sheet(
            h1_excel,
            description=(
                "Alle H1-Einzelantworten. Für die Lesbarkeit werden die Aufgaben "
                "innerhalb jedes Teilnehmers in einer festen Reihenfolge und die "
                "Visualisierungen als VSUP → ScaledGlyph → IsoGlyph dargestellt. "
                "Die tatsächlich präsentierte Reihenfolge bleibt in "
                "'Ursprüngliche Reihenfolge' erhalten."
            ),
            blank_between="participantId",
        ),

        "H1 Gepaarte Antworten": excel_sheet(
            paired_h1_excel,
            description=(
                "Nur Teilnehmer-Aufgaben, für die alle drei Visualisierungsmethoden "
                "gültig beantwortet wurden. Diese vollständigen Within-Subject-Paare "
                "bilden die Grundlage für die gepaarten Methodenvergleiche. "
                "Im Gegensatz zu 'H1 Einzelantworten' fehlen hier unvollständige "
                "Dreiergruppen."
            ),
            blank_between="participantId",
        ),

        "H1 Exakte Antworten": excel_sections(
            [
                (
                    "Verwendete Toleranzen",
                    exact_tol,
                    (
                        "Eine Antwort gilt rein deskriptiv als exakt, wenn ihre "
                        "absolute Abweichung höchstens 1 % der Wertspanne der "
                        "jeweiligen Variable beträgt."
                    ),
                ),
                (
                    "Exakte Antworten nach Aufgabe und Visualisierung",
                    exact_by_task,
                    (
                        "Diese Kennzahl ist nur eine Zusatzangabe und beeinflusst "
                        "ANOVA, t-Test, Friedman oder Wilcoxon nicht."
                    ),
                ),
            ]
        ),

        "H1 Teilnehmer Methoden": excel_sheet(
            pma,
            description=(
                "Auf Teilnehmer × Visualisierung aggregierte Werte ausschließlich aus "
                "vollständig gepaarten H1-Aufgaben. Unvollständige Dreiergruppen wurden vorher "
                "entfernt. Dadurch kann die hier ausgewiesene Anzahl berücksichtigter Aufgaben "
                "kleiner sein als in 'H1 Übersicht'. Diese Tabelle bildet die direkte Grundlage "
                "für Normalitätsprüfung, Gesamtvergleich und gegebenenfalls paarweise Vergleiche."
            ),
        ),

        "H1 Genauigkeit Statistik": excel_sections(
            [
                (
                    "1. Normalitätsprüfung",
                    normality,
                    "Prüfung der paarweisen Differenzen mit Shapiro-Wilk.",
                ),
                (
                    "2. Gesamtvergleich der Methoden",
                    omnibus,
                    (
                        "Bei erfüllter Normalitätsannahme: Repeated-Measures-ANOVA; "
                        "ansonsten Friedman-Test."
                    ),
                ),
                (
                    "3. Paarweise Vergleiche",
                    posthoc,
                    (
                        "Nur nach signifikantem Gesamtvergleich. Parametrisch: gepaarte "
                        "t-Tests; nichtparametrisch: Wilcoxon; jeweils Bonferroni-korrigiert."
                    ),
                ),
            ],
            description="Statistischer Entscheidungsweg für die H1-Genauigkeit.",
        ),

        "H1 Zeit Statistik": excel_sections(
            [
                ("1. Normalitätsprüfung", n_t, "Shapiro-Wilk der paarweisen Zeitdifferenzen."),
                (
                    "2. Gesamtvergleich der Methoden",
                    o_t,
                    "Repeated-Measures-ANOVA oder Friedman-Test für die Antwortzeit.",
                ),
                (
                    "3. Paarweise Vergleiche",
                    p_t,
                    "Gepaarte t-Tests oder Wilcoxon mit Bonferroni-Korrektur.",
                ),
            ],
            description="Sekundäre H1-Auswertung der Antwortzeit.",
        ),

        "H2 Übersicht": excel_sheet(summary),

        "H2 Regionsauswahl": excel_sheet(
            region_excel,
            description=(
                "Einzelentscheidungen der Regionsaufgabe inklusive ausgewählter Region, "
                "zugehöriger Werte, subjektiver Sicherheit, Antwortzeit und Begründung."
            ),
        ),

        "H2 Auswahlhäufigkeit": excel_sheet(freq),

        "H2 Unsicherheit Statistik": excel_sections(
            [
                ("1. Normalitätsprüfung", n_r, "Shapiro-Wilk der paarweisen Differenzen."),
                (
                    "2. Gesamtvergleich der Methoden",
                    o_r,
                    "Vergleich der Unsicherheit der gewählten Regionen zwischen den Methoden.",
                ),
                (
                    "3. Paarweise Vergleiche",
                    p_r,
                    "Paarweise Vergleiche nur nach signifikantem Gesamtvergleich.",
                ),
            ],
            description="Primäre inferenzstatistische H2-Auswertung.",
        ),

        "H2 Zielabweichung Statistik": excel_sections(
            [
                ("1. Normalitätsprüfung", n_d, "Shapiro-Wilk der paarweisen Differenzen."),
                (
                    "2. Gesamtvergleich der Methoden",
                    o_d,
                    "Zusätzlicher Vergleich der Abweichung vom Zielwert -12 °C.",
                ),
                (
                    "3. Paarweise Vergleiche",
                    p_d,
                    "Paarweise Vergleiche nur nach signifikantem Gesamtvergleich.",
                ),
            ],
            description=(
                "Zusätzliche H2-Analyse; die Unsicherheit der gewählten Region "
                "bleibt die primäre H2-Messgröße."
            ),
        ),

        "Feedback": excel_sheet(
            feedback_df,
            description=(
                "Nur das freie Textfeld aus dem Abschlussfeedback. "
            ),
        ),
    }

    export_results_xlsx(
        output_dir / "00_Studienauswertung.xlsx",
        workbook_sheets,
    )

    # Kurze Konsolenausgabe -------------------------------------------
    print(
        f"Analysierte Teilnehmende: "
        f"{base['participantId'].nunique()}"
    )

    print(
        f"H1-Klickaufgaben: {len(h1)}"
    )

    if not h1.empty:
        print(
            f"  beantwortet: "
            f"{int(h1['answered'].sum())}"
        )
        print(
            f"  nicht beantwortet (-999/NaN): "
            f"{int((~h1['answered']).sum())}"
        )

    print(
        f"H2-Regionsentscheidungen: {len(region)}"
    )

    if not region.empty:
        print("\nSicherste Region je Visualisierung:")
        for method in METHOD_ORDER:
            values = REGION_VALUES[method]
            safest_region = min(
                values,
                key=lambda r: values[r]["uncertainty"]
            )
            unc = values[safest_region]["uncertainty"]
            print(
                f"  {method}: {safest_region} "
                f"(Unsicherheit {unc})"
            )

    print(
        f"\nExakte Antworten: Toleranz = {EXACT_PERCENTAGE * 100:.1f} % "
        "der jeweiligen Variablenspanne (rein deskriptiv)."
    )
    print(
        "Testentscheidung: siehe analysis_output/analysis_decision.txt"
    )

    print(
        "\nFertig. Alle Ergebnisdateien liegen in:"
    )
    print(output_dir)
    print("Excel-Gesamtübersicht: 00_Studienauswertung.xlsx (deutsche, zusammengefasste Tabellenblätter)")
    print("CSV-Dateien: Semikolon-getrennt und UTF-8-BOM für deutsches Excel")


if __name__ == "__main__":
    main()